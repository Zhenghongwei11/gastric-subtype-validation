from __future__ import annotations

import argparse
import csv
from pathlib import Path

import openpyxl


SHEET_DATASET_IDS = {
    "KUGH": "GSE26899",
    "YUHS": "GSE13861",
    "KUCM": "GSE26901",
    "SMC": "SMC",
    "MDACC": "MDACC",
    "ACRG": "ACRG",
    "TCGA": "TCGA",
}

SHEET_ORDER = ["KUGH", "YUHS", "KUCM", "SMC", "MDACC", "ACRG", "TCGA"]

OUTPUT_FIELDS = [
    "dataset_id",
    "source_sheet",
    "row_id",
    "patient_id",
    "array_id",
    "geo_id",
    "sample_name",
    "subgroup",
    "sex",
    "age",
    "location",
    "lauren",
    "ajcc_stage",
    "m_stage",
    "pathologic_t",
    "pathologic_n",
    "pathologic_m",
    "overall_survival_event",
    "overall_survival_months",
    "recurrence_event",
    "recurrence_free_survival_months",
    "adjuvant_chemotherapy",
    "adjuvant_chemotherapy_binary",
    "chemotherapy",
    "chemotherapy_binary",
    "radiation_therapy",
    "tcga_subtype",
    "source_file",
]


def normalize_value(value: object) -> str:
    if value in (None, ""):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return format(value, ".10g")
    return str(value).strip()


def first_present(row: dict[str, str], field_names: list[str]) -> str:
    for field_name in field_names:
        value = row.get(field_name, "")
        if value:
            return value
    return ""


def normalize_binary(value: str) -> str:
    cleaned = value.strip().lower()
    if cleaned in {"1", "yes", "y", "true"}:
        return "1"
    if cleaned in {"0", "no", "n", "false"}:
        return "0"
    return ""


def write_tsv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()), delimiter="\t")
        writer.writeheader()
        writer.writerows(rows)


def parse_sheet_rows(workbook_path: Path, sheet_name: str) -> tuple[list[dict[str, str]], dict[str, str]]:
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    header_values = [normalize_value(value) for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]

    rows: list[dict[str, str]] = []
    for raw_values in sheet.iter_rows(min_row=2, values_only=True):
        values = [normalize_value(value) for value in raw_values]
        row = {
            header_values[index]: values[index]
            for index in range(min(len(header_values), len(values)))
            if header_values[index]
        }
        row_id = first_present(row, ["Patients_ID", "Array id", "SCRI No.", "barcode"])
        if not row_id:
            continue

        chemotherapy = first_present(row, ["Chemotherapy", "Adjuvant.chemo"])
        standardized_row = {
            "dataset_id": SHEET_DATASET_IDS[sheet_name],
            "source_sheet": sheet_name,
            "row_id": row_id,
            "patient_id": row_id,
            "array_id": row.get("Array id", ""),
            "geo_id": first_present(row, ["GEO ID", "GEO_ID"]),
            "sample_name": row.get("Sample\nName", ""),
            "subgroup": row.get("Subgroup", ""),
            "sex": first_present(row, ["Sex", "sex", "Gender"]),
            "age": first_present(row, ["Age", "age", "Age at Initial Diagnosis"]),
            "location": first_present(row, ["Location", "Anatomic Region", "revised location"]),
            "lauren": first_present(row, ["Lauren", "Lauren Class"]),
            "ajcc_stage": first_present(row, ["AJCC.stage", "AJCC6", "Stage", "TNM Stage", "stage(TNM)"]),
            "m_stage": first_present(row, ["M.stage", "M", "Pathologic M"]),
            "pathologic_t": first_present(row, ["T", "Pathologic T"]),
            "pathologic_n": first_present(row, ["N", "Pathologic N"]),
            "pathologic_m": first_present(row, ["M", "Pathologic M"]),
            "overall_survival_event": first_present(row, ["Death (1=yes, 0=no)", "Death"]),
            "overall_survival_months": row.get("OS.m", ""),
            "recurrence_event": first_present(row, ["Recurrence (1=yes, 0=no)", "Recur"]),
            "recurrence_free_survival_months": first_present(row, ["RFS.m", "DFS.m"]),
            "adjuvant_chemotherapy": row.get("Adjuvant.chemo", ""),
            "adjuvant_chemotherapy_binary": normalize_binary(row.get("Adjuvant.chemo", "")),
            "chemotherapy": chemotherapy,
            "chemotherapy_binary": normalize_binary(chemotherapy),
            "radiation_therapy": row.get("RadiationTherapy", ""),
            "tcga_subtype": row.get("TCGA Subtype", ""),
            "source_file": workbook_path.name,
        }
        rows.append(standardized_row)

    summary_row = {
        "dataset_id": SHEET_DATASET_IDS[sheet_name],
        "source_sheet": sheet_name,
        "row_count": str(len(rows)),
        "os_nonmissing_rows": str(sum(1 for row in rows if row["overall_survival_months"])),
        "os_event_nonmissing_rows": str(sum(1 for row in rows if row["overall_survival_event"])),
        "rfs_nonmissing_rows": str(sum(1 for row in rows if row["recurrence_free_survival_months"])),
        "recurrence_event_nonmissing_rows": str(sum(1 for row in rows if row["recurrence_event"])),
        "adjuvant_nonmissing_rows": str(sum(1 for row in rows if row["adjuvant_chemotherapy_binary"])),
        "source_file": workbook_path.name,
    }
    return rows, summary_row


def parse_rows(input_path: Path) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    clinical_rows: list[dict[str, str]] = []
    summary_rows: list[dict[str, str]] = []
    for sheet_name in SHEET_ORDER:
        sheet_rows, summary_row = parse_sheet_rows(input_path, sheet_name)
        clinical_rows.extend(sheet_rows)
        summary_rows.append(summary_row)
    return clinical_rows, summary_rows


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Standardize Nat Commun 2018 Supplementary Data 2 workbook")
    parser.add_argument("--input", required=True)
    parser.add_argument("--clinical-output", required=True)
    parser.add_argument("--summary-output", required=True)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    clinical_rows, summary_rows = parse_rows(Path(args.input))
    if not clinical_rows:
        raise SystemExit("No Nat Commun clinical rows were parsed")
    write_tsv(Path(args.clinical_output), clinical_rows)
    write_tsv(Path(args.summary_output), summary_rows)


if __name__ == "__main__":
    main()